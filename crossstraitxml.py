#!/usr/bin/env python
# -*- coding: utf-8 -*-

import io
import re
import time
import argparse
import zipfile

import opencc
from openpyxl import load_workbook
from lxml import etree as ET
from tqdm import tqdm
from data import abbreviations, simplified_separated, both_separated, tables

# Metadata
version = "1.0"
dictionaryname = u"兩岸詞典"
currenttime = time.strftime("%d-%m-%Y %H:%M:%S", time.localtime())
dtd_url = "https://raw.github.com/soshial/xdxf_makedict/master/format_standard/xdxf_strict.dtd"
doctypestring = "<!DOCTYPE xdxf SYSTEM \'%s\'>" % dtd_url
creation_date = time.strftime("%d-%m-%Y", time.localtime())
declaration = ("CrossStraitXML: Cross-strait Dictionary (XLSX format) to XDXF"
               " Converter\nVersion %s\n" % version)
src_url = "https://github.com/g0v/moedict-data-csld"
license = "CC BY-NC-ND 4.0"
description = (u"Cross-strait Chinese Dictionary (兩岸常用詞典) is a modern, "
               "good-quality Chinese dictionary made freely available by the "
               "Taiwanese Ministry of Education. It focuses  mainly on "
               "words that have different meanings or pronunciations in "
               "Taiwan (R.o.C) and Mainland China (P.R.C.), or which are "
               "specific to one of these regions. The dictionary was "
               "released under a CC BY-NC-ND 4.0 license._lb_This XDXF file "
               "was created automatically by the CrossStraitXML converter, "
               "version %s on %s._lb_CrossStraitXML converter is free and "
               "unencumbered software released into the public domain.") % (version,
                                                                            currenttime)
# Publishing date will be set later, the "modified date" on the XLSX metadata
# will be used for this. Dictionary date will be "1. + publishing date +
# converter version.
publishing_date = ""
dictionary_version = ""
xlsx_file = u"兩岸詞典.xlsx"
entries = list()
entry_list = list()


def simplify(trad_string):
    """Convert Taiwanese Traditional Chinese to Simplified Chinese.

    Converts a string in (Taiwan) traditional Chinese into simplified
    Chinese using OpenCC.
    """
    simplified = opencc.convert(trad_string, config="tw2s.json")
    return(simplified)


def multi_replace(inputstring, replacements):
    """Apply the replace method multiple times.

    "inputstring" is self-explanatory, "replacements" is a list of
    tuples, the fist item of each tuple the substring to be replaced
    and the second the replacement text.
    """
    for replacement in replacements:
        if replacement[0] in inputstring:
            inputstring = inputstring.replace(replacement[0], replacement[1])
    return inputstring


def multi_replace_final(inputlist, replacements):
    """Apply the replace method multiple times and join lines.

    "inputlist" is a list of strings, each to have replace method applied
    multiple times. "replacements" is a list of tuples, the fist item of each
    tuple the substring to be replaced and the second the replacement text.
    After all replacements are applied, the string will be joined with a new
    line character between each line and the resulting string will be returned.
    """
    inputlist = inputlist.split("\n")
    for line in inputlist:
        for replacement in replacements:
            if replacement[0] in line:
                line = line.replace(replacement[0], replacement[1])
    outputstring = "\n".join(inputlist)
    return outputstring


def excel_to_dict(worksheet):
    u"""Convert XLSX to a well-structured list of python dictionaries.

    Converts the Cross-strait Dictionary original XLSX file worksheet (as
    opened by openpyxl) and converts it into a list of python dictionaries to
    be used for writing the XDXF file. The keys of each dictionary are as
    follows:

    gaojianbanben: draft version
    gaojianjieduan: draft stage
    gaojianzhuangtai: draft status
    beizhu: (current draft related) remark
    ziciliushuixu: entry number
    zhengtizixing: entry in traditional characters
    jianhuazixing: entry in simplified characters (if different)
    yinxu: phonetic sequence (order according to pinyin when characters are
           the same but pronunciation differs)
    tailuteyouci: "★" if the entry is specific to Taiwan, "▲" if specific to
                  the mainland
    tailuteyouyin: "★" if the entry refers to a pronunciation specific to
                   Taiwan, "▲" if specific to the mainland
    taiwanyindu: Taiwan pronunciation in bopomofo
    taiwanhanpin: Taiwan pronunciation in pinyin
    daluyindu: Mainland pronunciation in bopomofo (if different)
    daluhanpin: Mainland pronunciation in pinyin (if different)
    definitions: definitions (itself a dictionary with numbered keys)
    examples: lists of example sentences (dictionary with numbered keys
              corresponding to the definitions)
    synonyms: lists of synonyms (dictionary with numbered keys corresponding
              to the definitions)
    """
    entry_list = list()
    rows = worksheet.max_row - 1
    for entry in tqdm(worksheet.iter_rows(row_offset=1),
                      desc="Processing entries", unit=" entries",
                      total=rows):
        if entry[4].value:  # Ignore empty lines.
            entry_dict = ({"gaojianbanben": entry[0].value,
                           "gaojianjieduan": entry[1].value,
                           "gaojianzhuangtai": entry[2].value,
                           "beizhu": entry[3].value,
                           "ziciliushuixu": entry[4].value,
                           "zhengtizixing": entry[5].value,
                           "jianhuazixing": entry[6].value,
                           "yinxu": entry[7].value,
                           "tailuteyouci": entry[8].value,
                           "tailuteyouyin": entry[9].value,
                           "taiwanyindu": entry[10].value,
                           "taiwanhanpin": entry[11].value,
                           "daluyindu": entry[12].value,
                           "daluhanpin": entry[13].value})
            # Replace separate radicals for the actual characters.
            entry_dict["jianhuazixing"] = multi_replace(entry_dict["jianhuazixing"],
                                                        simplified_separated)
            entry_dict["jianhuazixing"] = multi_replace(entry_dict["jianhuazixing"],
                                                        both_separated)
            entry_dict["zhengtizixing"] = multi_replace(entry_dict["zhengtizixing"],
                                                        both_separated)
            entry_dict["definitions"] = dict()
            entry_dict["examples"] = dict()
            entry_dict["synonyms"] = dict()
            for x in range(1, 30):
                if entry[x + 13].value:
                    entry_dict["definitions"][x] = entry[x + 13].value
                    # Separate definitions.
                    if len(re.findall(ur"^[0-9]+\.",
                           entry_dict["definitions"][x])) > 0:
                        entry_dict["definitions"][x] = re.findall(r"^[0-9]+\.(.*)",
                                                                  entry_dict["definitions"][x])[0]
                    # Substitute separate radicals with the correct
                    # characters.
                    entry_dict["definitions"][x] = multi_replace(entry_dict["definitions"][x],
                                                                 both_separated)
                    # Substitute HTML tables with plain text.
                    if u"</table>" in entry_dict["definitions"][x]:
                        entry_dict["definitions"][x] = multi_replace(entry_dict["definitions"][x],
                                                                     tables)
                    # Add link to "参见" references.
                    if len(re.findall(ur"參見【(.+?)】。",
                                      entry_dict["definitions"][x])) > 0:
                        entry_dict["definitions"][x] = re.sub(ur"參見【(.+?)】。",
                                                              ur"參見【_lt_kref_mt_\1_lt_/kref_mt_】。",
                                                              entry_dict["definitions"][x])
                    # Remove <font face> tags.
                    if len(re.findall(ur"<font face=\".+?\">(.+?)</font>",
                                      entry_dict["definitions"][x])) > 0:
                        entry_dict["definitions"][x] = re.sub(ur"<font face=\".+?\">(.+?)</font>",
                                                              ur"\1",
                                                              entry_dict["definitions"][x])
                    # Entry 罗马数字 has one X with css for overline. Convert
                    # it to X with unicode overline.
                    if u"""<span style="text-decoration:overline;">X</span>""" in entry_dict["definitions"][x]:
                        entry_dict["definitions"][x] = entry_dict["definitions"][x].replace(u"""<span style="text-decoration:overline;">X</span>""", u"X̅")
                    # Use plain text _underlined marks_ since proper underline
                    # is not supported.
                    if u"u>" in entry_dict["definitions"][x]:
                        entry_dict["definitions"][x] = multi_replace(entry_dict["definitions"][x],
                                                                     [("<u>", "_"), ("</u>", "_")])
                    # Mark regionalisms.
                    if entry_dict["tailuteyouci"] == u"★":
                        entry_dict["definitions"][x] = u"【陸】" + entry_dict["definitions"][x]
                    if entry_dict["tailuteyouci"] == u"▲":
                        entry_dict["definitions"][x] = u"【臺】" + entry_dict["definitions"][x]
                    entry_dict["definitions"][x] = multi_replace(entry_dict["definitions"][x],
                                                                 [(u"<sup>", u"_lt_sup_mt_"),
                                                                  (u"</sup>", u"_lt_/sup_mt_"),
                                                                  (u"<sub>", u"_lt_sub_mt_"),
                                                                  (u"</sub>", u"_lt_/sub_mt_"),
                                                                  (u"▲", u"【臺】"),
                                                                  (u"★", u"【陸】")])
                    # Mark abbreviations.
                    for abbreviation in abbreviations:
                        entry_dict["definitions"][x] = entry_dict["definitions"][x].replace(abbreviation[0],
                                                                                            "_lt_abbr_mt_%s_lt_/abbr_mt_" % abbreviation[0])
                    # Convert to simplified Chinese (if selected).
                    if args.simplified:
                        entry_dict["definitions"][x] = simplify(entry_dict["definitions"][x])
                    # Extract synonyms.
                    if len(re.findall(u"(∥也作「.+」。)",
                                      entry_dict["definitions"][x])) > 0:
                        entry_dict["synonyms"][x] = re.findall(u"∥也作「(.+)」。",
                                                               entry_dict["definitions"][x])[0]
                        entry_dict["definitions"][x] = entry_dict["definitions"][x].replace(re.findall(u"(∥也作「.+」。)",
                                                                                            entry_dict["definitions"][x])[0], "")
                        entry_dict["synonyms"][x] = multi_replace(entry_dict["synonyms"][x],
                                                                  [(u"「", ""), (u"」", "")])
                        entry_dict["synonyms"][x] = entry_dict["synonyms"][x].split(u"、")
                    # Extract examples.
                    if len(re.findall(u"\[例\](.+)[。？！]",
                                      entry_dict["definitions"][x])) > 0:
                        entry_dict["examples"][x] = re.findall(u"\[例\](.+[。？！])",
                                                               entry_dict["definitions"][x])[0]
                        entry_dict["definitions"][x] = entry_dict["definitions"][x].replace(re.findall(u"(\[例\].+[。？！])",
                                                                                                       entry_dict["definitions"][x])[0], "")
                        entry_dict["examples"][x] = multi_replace(entry_dict["examples"][x],
                                                                  [(u"│", u"｜"),
                                                                   (u"︱", u"｜"),
                                                                   (u"∣", u"｜")]).split(u"｜")
            entry_list.append(entry_dict)
    for entry_dict in entry_list:    # Remove empty keys.
        for key in entry_dict.keys():
            if not entry_dict[key]:
                entry_dict.pop(key, None)
    return(entry_list)


def set_dates(xlsx_file):
    """Set variables publishing_date and dictionary_version.

    Read XLSX file metadata and use the "date modified" for the variables
    publishing_date and dictionary_version.
    """
    zf = zipfile.ZipFile(xlsx_file)
    core = zf.read("docProps/core.xml")
    date_modified = re.findall("<dcterms:modified xsi:type=\"dcterms:"
                               "W3CDTF\">([0-9-]+?)T", core)[0]
    date = (date_modified[8:] + "-" + date_modified[5:7] + "-" +
            date_modified[:4])
    global publishing_date
    publishing_date = date
    global dictionary_version
    dictionary_version = "1." + date_modified.replace("-", "") + "-" + version


def create_xdxf(entry_list):
    """Create an XML object consisting of the final XDXF file.

    Create an ET XML object containing all the XDXF data from the list of
    dictionaries described above.
    """
    # Metadata
    xdxfdic_top = ET.Element("xdxf", lang_from="CHI", lang_to="CHI",
                             format="logical", revision="33")
    meta_info = ET.SubElement(xdxfdic_top, "meta_info")
    lexicon = ET.SubElement(xdxfdic_top, "lexicon")
    if args.simplified:
        meta_info_title = ET.SubElement(meta_info,
                                        "title").text = simplify(dictionaryname)
        meta_info_full_title = ET.SubElement(meta_info,
                                             "full_title").text = simplify(dictionaryname)
    else:
        meta_info_title = ET.SubElement(meta_info,
                                        "title").text = dictionaryname
        meta_info_full_title = ET.SubElement(meta_info,
                                             "full_title").text = dictionaryname
    if args.simplified:
        meta_info_publisher = ET.SubElement(meta_info, "publisher").text = u"中华民国教育部 (Taiwan Ministry of Education)"
    else:
        meta_info_publisher = ET.SubElement(meta_info, "publisher").text = u"中華民國教育部 (Taiwan Ministry of Education)"
    meta_info_description = ET.SubElement(meta_info,
                                          "description").text = description
    # Abbreviations
    meta_info_abbreviations = ET.SubElement(meta_info, "abbreviations")
    abbrlist = []
    for tupple in abbreviations:
        abbrlist.append(tupple[0])
    for abbreviation in abbreviations:
        # Convert abbreviations to simplified Chinese, if selected.
        if args.simplified:
            simplified_abbreviation = list()
            for part in abbreviation:
                simplified_abbreviation.append(simplify(part))
            abbreviation = (simplified_abbreviation[0],
                            simplified_abbreviation[1],
                            simplified_abbreviation[2])
        if abbreviation[2] is not "":
            current_abbr_def = ET.SubElement(meta_info_abbreviations,
                                             "abbr_def", type=abbreviation[2])
            abbr_k = ET.SubElement(current_abbr_def,
                                   "abbr_k").text = abbreviation[0]
            abbr_v = ET.SubElement(current_abbr_def,
                                   "abbr_v").text = (abbreviation[1])
        else:
            current_abbr_def = ET.SubElement(meta_info_abbreviations,
                                             "abbr_def")
            abbr_k = ET.SubElement(current_abbr_def,
                                   "abbr_k").text = abbreviation[0]
            abbr_v = ET.SubElement(current_abbr_def,
                                   "abbr_v").text = abbreviation[1]
    # More metadata
    meta_info_file_ver = ET.SubElement(meta_info,
                                       "file_ver").text = (dictionary_version)
    meta_info_creation_date = ET.SubElement(meta_info,
                                            "creation_date").text = (creation_date)
    meta_info_publishing_date = (ET.SubElement
                                 (meta_info,
                                  "publishing_date").text) = (publishing_date)
    meta_info_dict_src_url = ET.SubElement(meta_info,
                                           "dict_src_url").text = src_url
    # Process dictionary entries.
    for entry in tqdm(entry_list, desc="Writing XDXF", unit=" entries"):
        lexicon_ar = ET.SubElement(lexicon, "ar")
        if "jianhuazixing" in entry:
            lexicon_ar_k = ET.SubElement(lexicon_ar,
                                         "k").text = entry["jianhuazixing"]
        else:
            lexicon_ar_k = ET.SubElement(lexicon_ar,
                                         "k").text = entry["zhengtizixing"]
        lexicon_ar_k_trad = ET.SubElement(lexicon_ar,
                                          "k").text = entry["zhengtizixing"]
        lexicon_ar_def = ET.SubElement(lexicon_ar, "def")
        lexicon_ar_def_grtr = ET.SubElement(lexicon_ar_def, "gr")
        if ("taiwanhanpin" in entry and "daluhanpin" not in entry) or ("daluhanpin" in entry and "taiwanhanpin" not in entry):
            if "taiwanhanpin" in entry:
                lexicon_ar_def_grtr_tr = ET.SubElement(lexicon_ar_def_grtr,
                                                       "tr").text = entry["taiwanhanpin"]
            if "daluhanpin" in entry:
                lexicon_ar_def_grtr_tr = ET.SubElement(lexicon_ar_def_grtr,
                                                       "tr").text = entry["daluhanpin"]
        if "taiwanhanpin" in entry and "daluhanpin" in entry:
            if entry["taiwanhanpin"] == entry["daluhanpin"]:
                lexicon_ar_def_grtr_tr = ET.SubElement(lexicon_ar_def_grtr,
                                                       "tr").text = entry["taiwanhanpin"]
            else:
                if args.simplified:
                    lexicon_ar_def_grtr_tr1 = ET.SubElement(lexicon_ar_def_grtr,
                                                            "tr").text = u"【陆】" + entry["daluhanpin"]
                    lexicon_ar_def_grtr_tr2 = ET.SubElement(lexicon_ar_def_grtr,
                                                            "tr").text = u"【台】" + entry["taiwanhanpin"]
                else:
                    lexicon_ar_def_grtr_tr1 = ET.SubElement(lexicon_ar_def_grtr,
                                                            "tr").text = u"【陸】" + entry["daluhanpin"]
                    lexicon_ar_def_grtr_tr2 = ET.SubElement(lexicon_ar_def_grtr,
                                                            "tr").text = u"【臺】" + entry["taiwanhanpin"]
        if args.bopomofo:
            if ("taiwanyindu" in entry and "daluyindu" not in entry) or ("daluyindu" in entry and "taiwanyindu" not in entry):
                if "taiwanyindu" in entry:
                    lexicon_ar_def_grtr_trb = ET.SubElement(lexicon_ar_def_grtr,
                                                            "tr").text = entry["taiwanyindu"]
                if "daluyindu" in entry:
                    lexicon_ar_def_grtr_trb = ET.SubElement(lexicon_ar_def_grtr,
                                                            "tr").text = entry["daluyindu"]
            elif entry["taiwanyindu"] == entry["daluyindu"]:
                lexicon_ar_def_grtr_trb = ET.SubElement(lexicon_ar_def_grtr,
                                                        "tr").text = entry["taiwanyindu"]
            elif "taiwanyindu" in entry and "daluyindu" in entry:
                if args.simplified:
                    lexicon_ar_def_grtr_trb1 = ET.SubElement(lexicon_ar_def_grtr,
                                                             "tr").text = u"【陆】" + entry["daluyindu"]
                    lexicon_ar_def_grtr_trb2 = ET.SubElement(lexicon_ar_def_grtr,
                                                             "tr").text = u"【台】" + entry["taiwanyindu"]
                else:
                    lexicon_ar_def_grtr_trb1 = ET.SubElement(lexicon_ar_def_grtr,
                                                             "tr").text = u"【陸】" + entry["daluyindu"]
                    lexicon_ar_def_grtr_trb2 = ET.SubElement(lexicon_ar_def_grtr,
                                                             "tr").text = u"【臺】" + entry["taiwanyindu"]
        for i in range(1, 31):
            if i in entry["definitions"]:
                globals()['lexicon_ar_def_def%s' % i] = ET.SubElement(lexicon_ar_def, "def")
                if args.simplified:
                    deftext = simplify(entry["definitions"][i])
                else:
                    deftext = entry["definitions"][i]
                globals()['lexicon_ar_def_def%s_deftext' % i] = (ET.SubElement(globals()['lexicon_ar_def_def%s' % i],
                                                                 "deftext").text) = deftext
                if "synonyms" in entry and i in entry["synonyms"]:
                        globals()['lexicon_ar_def_def%s_sr' % (i)] = ET.SubElement(globals()['lexicon_ar_def_def%s' % i],
                                                                                   "sr")
                        for y in range(0, len(entry["synonyms"][i])):
                            if args.simplified:
                                globals()['lexicon_ar_def_def%s_sr_syn%s' % (i, y)] = (ET.SubElement(globals()['lexicon_ar_def_def%s_sr' % (i)],
                                                                                                     "kref", type="syn").text) = simplify(entry["synonyms"][i][y])
                            else:
                                globals()['lexicon_ar_def_def%s_sr_syn%s' % (i, y)] = (ET.SubElement(globals()['lexicon_ar_def_def%s_sr' % (i)],
                                                                                                     "kref", type="syn").text) = entry["synonyms"][i][y]
                if "examples" in entry and i in entry["examples"]:
                    for y in range(0, len(entry["examples"][i])):
                        globals()['lexicon_ar_def_def%s_ex%s' % (i, y)] = ET.SubElement(globals()['lexicon_ar_def_def%s' % i],
                                                                                        "ex", type="exm")
                        if args.simplified:
                            globals()['lexicon_ar_def_def%s_ex%s_exorig' % (i, y)] = (ET.SubElement(globals()['lexicon_ar_def_def%s_ex%s' % (i, y)],
                                                                                      "exorig").text) = simplify(entry["examples"][i][y])
                        else:
                            globals()['lexicon_ar_def_def%s_ex%s_exorig' % (i, y)] = (ET.SubElement(globals()['lexicon_ar_def_def%s_ex%s' % (i, y)],
                                                                                      "exorig").text) = entry["examples"][i][y]
    return xdxfdic_top


argparser = argparse.ArgumentParser()
argparser.add_argument("-i", "--input_file",
                       help="Original Cross-strait Dictionary in XLSX "
                       "format. (By default it is named \"兩岸詞典.xlsx\".) "
                       "When this argument is not provided the program will "
                       "try to load a file named \"兩岸詞典.xlsx\" in the "
                       "current folder.")
argparser.add_argument("-s", "--simplified", help="The definitions' text will"
                       "be automatically converted to Simplified Chinese "
                       "using OpenCC. Automatic conversion is not perfect, "
                       "resulting text will contain errors. Headwords come "
                       "from the original XLSX file, they are not converted "
                       "and are always included in both variants.",
                       action="store_true")
argparser.add_argument("-b", "--bopomofo", help="Include bopomofo "
                       "transcriptions. Pinyin transcriptions are always "
                       "included.", action="store_true")
argparser.add_argument("-o", "--output_file", help="Name and path to "
                       "resulting XDXF file. If not provided the file will be "
                       "named \"兩岸詞典_ + dictionary version + .xdxf\" and "
                       "written to the current folder.")
args = argparser.parse_args()
if args.input_file:
    xlsx_file = args.input_file
print declaration
# Try to load the file and worksheet. Exit if it doesn't seem the right file.
print u"Loading the file…"
try:
    wb = load_workbook(filename=xlsx_file)
    ws = wb[u"工作表1"]
except:
    print ("Either the file doesn't exist or it doesn't seem to be the "
           "LAC original XLSX file.")
    exit()
print u"Converting…"
entry_list = excel_to_dict(ws)
set_dates(xlsx_file)
converted_xdxf = create_xdxf(entry_list)
xdxf_string = ET.tostring(converted_xdxf, encoding="utf-8", pretty_print=True,
                          xml_declaration=True,
                          doctype=doctypestring).decode("utf-8")
xdxf_string = multi_replace(xdxf_string, [("_lb_", "<br/>"), ("_lt_", "<"),
                            ("_mt_", ">")])
if args.output_file:
    xdxf_filename = args.output_file
else:
    if args.simplified:
        xdxf_filename = u"两岸词典_" + dictionary_version + ".xdxf"
    else:
        xdxf_filename = u"兩岸詞典_" + dictionary_version + ".xdxf"
io.open(xdxf_filename, "w", encoding="utf8").write(xdxf_string)
print "\nSuccess! The Cross-strait Dictionary file was converted to \"%s\"." % xdxf_filename
